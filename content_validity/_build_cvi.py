"""
Build the CVI Analysis spreadsheet for AwadhVaani content validity review.

Produces 06_cvi_analysis.xlsx with:
  • Instructions tab — rubric, scale, formulas explained
  • Items & Ratings tab — all 195 items, 7 expert columns, auto-computed CVI/Mean/SD/Pc/κ*/Status
  • Summary tab — S-CVI/Ave · S-CVI/UA · mean κ* · per-type breakdown
  • Expert Panel tab — table to record reviewer credentials

Run from project root:  python content_validity/_build_cvi.py
"""
import json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.workbook.defined_name import DefinedName

ROOT = Path(__file__).resolve().parent.parent
OUT  = Path(__file__).resolve().parent / "06_cvi_analysis.xlsx"

# ── Load source data ────────────────────────────────────────────────────
vocab   = json.loads((ROOT / "src/data/vocabulary.json").read_text(encoding='utf-8'))
lessons = json.loads((ROOT / "src/data/lessons.json").read_text(encoding='utf-8'))
quiz    = json.loads((ROOT / "src/data/quiz.json").read_text(encoding='utf-8'))
stories = json.loads((ROOT / "src/data/stories.json").read_text(encoding='utf-8'))

items = []
v = l = q = s = 0

for w in vocab:
    v += 1
    items.append({'id': f'V{v:03d}', 'type': 'Vocabulary',
                  'awadhi': w['awadhi'], 'roman': w['roman'], 'english': w['eng'],
                  'category': w['cat'], 'context': '', 'source': 'vocabulary.json'})

for unit in lessons:
    for lesson in unit['lessons']:
        for sent in lesson['sentences']:
            l += 1
            items.append({'id': f'L{l:03d}', 'type': 'Lesson Sentence',
                          'awadhi': sent['awadhi'], 'roman': sent['roman'], 'english': sent['english'],
                          'category': unit['title'],
                          'context': f"Unit {unit['id']}: {unit['title']} / {lesson['name']}",
                          'source': 'lessons.json'})

for qst in quiz:
    q += 1
    correct = next((o for o in qst['options'] if o.get('correct')), None)
    items.append({'id': f'Q{q:03d}', 'type': 'Quiz Item',
                  'awadhi': qst['awadhi'], 'roman': qst.get('roman', ''),
                  'english': correct['text'] if correct else '',
                  'category': qst['type'],
                  'context': f"Prompt: {qst['prompt']} | Options: {' / '.join(o['text'] for o in qst['options'])}",
                  'source': 'quiz.json'})

for story in stories:
    for i, p in enumerate(story.get('paragraphs', [])):
        s += 1
        items.append({'id': f'S{s:03d}', 'type': 'Story Paragraph',
                      'awadhi': p['awadhi'], 'roman': p.get('roman', ''), 'english': p['english'],
                      'category': story.get('category', ''),
                      'context': f"{story['title']} (¶{i+1})", 'source': 'stories.json'})
    for vc in story.get('vocabulary', []):
        s += 1
        items.append({'id': f'S{s:03d}', 'type': 'Story Vocab',
                      'awadhi': vc['awadhi'], 'roman': vc['roman'], 'english': vc['english'],
                      'category': story.get('category', ''),
                      'context': f'Vocab from "{story["title"]}"', 'source': 'stories.json'})

print(f"Loaded {len(items)} items: {v} vocab · {l} lessons · {q} quiz · {s} story")

# ── Style helpers ───────────────────────────────────────────────────────
FONT = "Arial"
THIN = Side(border_style='thin', color='CCCCCC')
BORDER = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)

def cell(ws, coord, value, *, bold=False, size=10, color='000000', bg=None, h='left', v='center', wrap=False, border=False, fmt=None):
    c = ws[coord] if isinstance(coord, str) else ws.cell(*coord)
    c.value = value
    c.font = Font(name=FONT, size=size, bold=bold, color=color)
    c.alignment = Alignment(horizontal=h, vertical=v, wrap_text=wrap)
    if bg:     c.fill = PatternFill('solid', start_color=bg)
    if border: c.border = BORDER
    if fmt:    c.number_format = fmt
    return c

# ── Build workbook ──────────────────────────────────────────────────────
wb = Workbook()

# ═══════════════════════════════════════════════════════════════════════
# Sheet 1 — Instructions
# ═══════════════════════════════════════════════════════════════════════
inst = wb.active
inst.title = "Instructions"

inst.column_dimensions['A'].width = 22
inst.column_dimensions['B'].width = 90

cell(inst, 'A1', 'AwadhVaani — CVI Analysis Spreadsheet', bold=True, size=18, color='2D5A3D')
cell(inst, 'A2', 'Content Validity Index analysis for the Awadhi language-learning app',
     size=11, color='666666')
inst.row_dimensions[1].height = 28

# Section: Rubric
cell(inst, 'A4', 'THE 4-POINT RELEVANCE SCALE', bold=True, size=12, color='E27431')
rubric = [
    ('1', 'Not relevant — does not belong in an Awadhi learning curriculum'),
    ('2', 'Somewhat relevant — peripheral, could be omitted'),
    ('3', 'Quite relevant — should be included'),
    ('4', 'Highly relevant — essential to the curriculum'),
]
for i, (n, desc) in enumerate(rubric, start=5):
    cell(inst, f'A{i}', n, bold=True, size=11, h='center', bg='F8F4ED', border=True)
    cell(inst, f'B{i}', desc, size=11, border=True, wrap=True)

# Section: How CVI is computed
row = 10
cell(inst, f'A{row}', 'HOW THE METRICS ARE COMPUTED', bold=True, size=12, color='E27431')
row += 1
formulas = [
    ('I-CVI', 'Item-level CVI = (# experts rating ≥3) / (total # experts)'),
    ('S-CVI/Ave', 'Scale-level CVI (average) = mean of all I-CVI values'),
    ('S-CVI/UA', 'Universal Agreement CVI = (# items rated ≥3 by ALL experts) / total items'),
    ('Mean', 'Average of all expert ratings for the item (informational, not part of CVI)'),
    ('SD', 'Standard deviation of expert ratings — high SD = experts disagree'),
    ('Pc', 'Chance agreement: COMBIN(N,A) × 0.5^N · where N=experts, A=# rating ≥3'),
    ('κ* (Kappa)', 'Modified Kappa = (I-CVI − Pc) / (1 − Pc) · corrects CVI for chance agreement'),
]
for name, desc in formulas:
    cell(inst, f'A{row}', name, bold=True, size=11, h='left', bg='F8F4ED', border=True)
    cell(inst, f'B{row}', desc, size=11, border=True, wrap=True)
    row += 1

# Section: Thresholds
row += 1
cell(inst, f'A{row}', 'ACCEPTANCE THRESHOLDS', bold=True, size=12, color='E27431')
row += 1
thresholds = [
    ('I-CVI ≥ 0.78', 'PASS — for a panel of 6–10 experts (Lynn, 1986)'),
    ('I-CVI ≥ 0.83', 'PASS — for a panel of 3–5 experts (stricter, smaller N)'),
    ('I-CVI 0.60–0.78', 'BORDERLINE — needs revision or expert discussion'),
    ('I-CVI < 0.60', 'FAIL — revise or drop from curriculum'),
    ('S-CVI/Ave ≥ 0.90', 'Excellent — scale is content-valid'),
    ('S-CVI/Ave 0.80–0.89', 'Adequate — minor revisions needed'),
    ('S-CVI/UA ≥ 0.80', 'Strict universal-agreement target'),
    ('κ* > 0.74', 'Excellent agreement (gold standard for publication)'),
    ('κ* 0.60–0.74', 'Good'),
    ('κ* < 0.40', 'Poor — revise item'),
]
for name, desc in thresholds:
    cell(inst, f'A{row}', name, bold=True, size=10, h='left', bg='F8F4ED', border=True)
    cell(inst, f'B{row}', desc, size=10, border=True, wrap=True)
    row += 1

# Section: How to use
row += 1
cell(inst, f'A{row}', 'HOW TO USE THIS WORKBOOK', bold=True, size=12, color='E27431')
row += 1
steps = [
    'Step 1', 'On the "Expert Panel" tab, fill in details for E1–E7 (name, affiliation, background, experience).',
    'Step 2', 'On the "Items & Ratings" tab, enter each expert\'s rating (1–4) in their column (H–N) for every item they evaluated.',
    'Step 3', 'I-CVI, Mean, SD, Pc, κ*, and Status auto-compute as ratings are entered.',
    'Step 4', 'The "Summary" tab updates automatically: S-CVI/Ave, S-CVI/UA, mean κ*, pass/borderline/fail counts, per-type breakdown.',
    'Step 5', 'Items marked FAIL or BORDERLINE go in your revision queue. Item-level κ* < 0.40 should be revised or dropped.',
    'Step 6', 'Use the "Summary" numbers as the headline result in your validity report.',
]
for i in range(0, len(steps), 2):
    cell(inst, f'A{row}', steps[i],   bold=True, size=10, bg='F8F4ED', border=True, h='center')
    cell(inst, f'B{row}', steps[i+1], size=10, border=True, wrap=True)
    row += 1

# Color legend
row += 1
cell(inst, f'A{row}', 'COLOR LEGEND (in the Items tab)', bold=True, size=12, color='E27431')
row += 1
legends = [
    ('PASS',       'C6F6D5', '008000', 'I-CVI ≥ 0.78 — item is content-valid'),
    ('BORDERLINE', 'FFF3CD', 'C77800', 'I-CVI 0.60–0.78 — needs discussion'),
    ('FAIL',       'FFD6D6', 'CC0000', 'I-CVI < 0.60 — revise or drop'),
    ('Awaiting',   'E8E8E8', '666666', 'No expert ratings entered yet'),
]
for label, bg, fg, desc in legends:
    cell(inst, f'A{row}', label, bold=True, size=10, h='center', bg=bg, color=fg, border=True)
    cell(inst, f'B{row}', desc, size=10, border=True, wrap=True)
    row += 1

# ═══════════════════════════════════════════════════════════════════════
# Sheet 4 — Expert Panel  (build first so its values exist when Items refs them)
# ═══════════════════════════════════════════════════════════════════════
panel = wb.create_sheet("Expert Panel")

panel_headers = ['Expert ID', 'Name', 'Affiliation', 'Background',
                 'Years of Awadhi Experience', 'Region / Sub-dialect Familiarity', 'Email']
widths = [10, 22, 28, 22, 16, 28, 28]
for i, h in enumerate(panel_headers, start=1):
    cell(panel, (1, i), h, bold=True, size=11, color='FFFFFF', bg='2D5A3D',
         h='center', border=True)
    panel.column_dimensions[get_column_letter(i)].width = widths[i-1]

panel.row_dimensions[1].height = 26

# 7 empty rows for experts
backgrounds = ['Native Speaker', 'Linguist', 'Educator', 'Cultural Scholar',
               'Native Speaker', 'Linguist', 'Educator']  # suggestion only
for i in range(1, 8):
    row = i + 1
    cell(panel, (row, 1), f'E{i}', bold=True, size=11, h='center', bg='F8F4ED', border=True)
    for col in range(2, 8):
        cell(panel, (row, col), '', size=11, border=True)
    panel.row_dimensions[row].height = 26

# Parameter box (used by Items & Summary formulas)
cell(panel, 'A11', 'PARAMETERS', bold=True, size=12, color='E27431')
cell(panel, 'A12', 'Total experts (N)',     bold=True, size=11, bg='F8F4ED', border=True)
cell(panel, 'B12', 7,                       size=11, bg='FFFF99', border=True, h='center')
cell(panel, 'A13', 'I-CVI pass threshold',  bold=True, size=11, bg='F8F4ED', border=True)
cell(panel, 'B13', 0.78,                    size=11, bg='FFFF99', border=True, h='center', fmt='0.00')
cell(panel, 'A14', 'I-CVI borderline floor',bold=True, size=11, bg='F8F4ED', border=True)
cell(panel, 'B14', 0.60,                    size=11, bg='FFFF99', border=True, h='center', fmt='0.00')

# Named ranges for formulas
wb.defined_names['N_EXPERTS'] = DefinedName(name='N_EXPERTS', attr_text="'Expert Panel'!$B$12")
wb.defined_names['CVI_PASS']  = DefinedName(name='CVI_PASS',  attr_text="'Expert Panel'!$B$13")
wb.defined_names['CVI_BORDER']= DefinedName(name='CVI_BORDER',attr_text="'Expert Panel'!$B$14")

# ═══════════════════════════════════════════════════════════════════════
# Sheet 2 — Items & Ratings
# ═══════════════════════════════════════════════════════════════════════
rt = wb.create_sheet("Items & Ratings", index=1)

# Header
headers = ['Item ID', 'Type', 'Awadhi', 'Roman', 'English', 'Category', 'Context',
           'E1', 'E2', 'E3', 'E4', 'E5', 'E6', 'E7',
           '# ≥3', 'I-CVI', 'Mean', 'SD', 'Pc', 'κ*', 'Status']
widths   = [9,       16,    18,       16,      26,        18,         34,
            6, 6, 6, 6, 6, 6, 6,
            7,       8,       7,      7,    8,    8,    14]

for i, h in enumerate(headers, start=1):
    bg = '2D5A3D' if i <= 7 else ('E27431' if 8 <= i <= 14 else '8B5C2E')
    cell(rt, (1, i), h, bold=True, size=10, color='FFFFFF', bg=bg, h='center', border=True)
    rt.column_dimensions[get_column_letter(i)].width = widths[i-1]
rt.row_dimensions[1].height = 30

# Freeze panes — keep header + item ID column visible
rt.freeze_panes = 'H2'

# Item rows + computed formulas
for idx, it in enumerate(items, start=2):
    cell(rt, (idx, 1), it['id'],       bold=True, size=10, h='center', border=True, bg='F8F4ED')
    cell(rt, (idx, 2), it['type'],     size=10, border=True)
    cell(rt, (idx, 3), it['awadhi'],   size=12, border=True)
    cell(rt, (idx, 4), it['roman'],    size=10, border=True, color='E27431')
    cell(rt, (idx, 5), it['english'],  size=10, border=True, color='666666')
    cell(rt, (idx, 6), it['category'], size=10, border=True)
    cell(rt, (idx, 7), it['context'],  size=9, border=True, wrap=True, color='888888')

    # E1–E7 empty rating cells (column H = 8 through N = 14)
    for col in range(8, 15):
        cell(rt, (idx, col), '', size=10, h='center', border=True, bg='FFFDF0')

    rng = f'H{idx}:N{idx}'   # E1 ... E7

    # # ≥3
    cell(rt, (idx, 15), f'=COUNTIF({rng},">=3")',
         size=10, h='center', border=True)

    # I-CVI
    cell(rt, (idx, 16),
         f'=IF(COUNT({rng})=0,"",COUNTIF({rng},">=3")/COUNT({rng}))',
         size=10, h='center', border=True, bold=True, fmt='0.00')

    # Mean
    cell(rt, (idx, 17),
         f'=IF(COUNT({rng})=0,"",AVERAGE({rng}))',
         size=10, h='center', border=True, fmt='0.00')

    # SD
    cell(rt, (idx, 18),
         f'=IF(COUNT({rng})<2,"",STDEV({rng}))',
         size=10, h='center', border=True, fmt='0.00')

    # Pc — COMBIN(N,A) × 0.5^N
    cell(rt, (idx, 19),
         f'=IF(COUNT({rng})=0,"",COMBIN(COUNT({rng}),COUNTIF({rng},">=3"))*0.5^COUNT({rng}))',
         size=10, h='center', border=True, fmt='0.000')

    # κ*  = (I-CVI − Pc) / (1 − Pc)
    cell(rt, (idx, 20),
         f'=IFERROR(IF(P{idx}="","",(P{idx}-S{idx})/(1-S{idx})),"")',
         size=10, h='center', border=True, fmt='0.00')

    # Status — uses absolute named-range thresholds
    cell(rt, (idx, 21),
         f'=IF(P{idx}="","Awaiting",IF(P{idx}>=CVI_PASS,"PASS",IF(P{idx}>=CVI_BORDER,"BORDERLINE","FAIL")))',
         bold=True, size=10, h='center', border=True)

# Conditional formatting for Status column (U)
status_range = f'U2:U{len(items)+1}'

green = PatternFill('solid', start_color='C6F6D5')
amber = PatternFill('solid', start_color='FFF3CD')
red   = PatternFill('solid', start_color='FFD6D6')
grey  = PatternFill('solid', start_color='E8E8E8')

rt.conditional_formatting.add(status_range, FormulaRule(formula=[f'$U2="PASS"'],       fill=green, font=Font(name=FONT, bold=True, color='008000')))
rt.conditional_formatting.add(status_range, FormulaRule(formula=[f'$U2="BORDERLINE"'], fill=amber, font=Font(name=FONT, bold=True, color='C77800')))
rt.conditional_formatting.add(status_range, FormulaRule(formula=[f'$U2="FAIL"'],       fill=red,   font=Font(name=FONT, bold=True, color='CC0000')))
rt.conditional_formatting.add(status_range, FormulaRule(formula=[f'$U2="Awaiting"'],   fill=grey,  font=Font(name=FONT, color='666666', italic=True)))

# Conditional formatting for I-CVI column (P)
icvi_range = f'P2:P{len(items)+1}'
rt.conditional_formatting.add(icvi_range, FormulaRule(formula=[f'AND(P2<>"",P2>=CVI_PASS)'],                            fill=green))
rt.conditional_formatting.add(icvi_range, FormulaRule(formula=[f'AND(P2<>"",P2>=CVI_BORDER,P2<CVI_PASS)'], fill=amber))
rt.conditional_formatting.add(icvi_range, FormulaRule(formula=[f'AND(P2<>"",P2<CVI_BORDER)'],                           fill=red))

# Data validation on rating cells (1-4)
from openpyxl.worksheet.datavalidation import DataValidation
dv = DataValidation(type='whole', operator='between', formula1=1, formula2=4,
                    allow_blank=True, showErrorMessage=True,
                    errorTitle='Invalid rating',
                    error='Please enter 1, 2, 3, or 4 (or leave blank if you did not rate this item)')
dv.add(f'H2:N{len(items)+1}')
rt.add_data_validation(dv)

# ═══════════════════════════════════════════════════════════════════════
# Sheet 3 — Summary
# ═══════════════════════════════════════════════════════════════════════
sm = wb.create_sheet("Summary", index=2)

sm.column_dimensions['A'].width = 38
sm.column_dimensions['B'].width = 14
sm.column_dimensions['C'].width = 38
sm.column_dimensions['D'].width = 14

cell(sm, 'A1', 'AwadhVaani — CVI Summary',
     bold=True, size=18, color='2D5A3D')
sm.row_dimensions[1].height = 30
cell(sm, 'A2', 'Headline metrics auto-compute as expert ratings are entered',
     size=10, color='666666')

# ── Headline metrics ────────────────────────────────────────────────
total_items = len(items)
items_range_p = f"'Items & Ratings'!P2:P{total_items+1}"   # I-CVI
items_range_t = f"'Items & Ratings'!T2:T{total_items+1}"   # κ*
items_range_b = f"'Items & Ratings'!B2:B{total_items+1}"   # Type
items_range_u = f"'Items & Ratings'!U2:U{total_items+1}"   # Status

metrics = [
    ('Total items in inventory',                f'=COUNTA(\'Items & Ratings\'!A2:A{total_items+1})', '0'),
    ('Items with at least one rating',         f'=COUNT({items_range_p})', '0'),
    ('Items with no ratings yet',              f'=COUNTIF({items_range_u},"Awaiting")', '0'),
    ('',                                        None, None),
    ('S-CVI / Ave  (target ≥ 0.90)',           f'=IFERROR(AVERAGE({items_range_p}),"—")',                       '0.000'),
    ('S-CVI / UA  (target ≥ 0.80)',            f'=IF(COUNT({items_range_p})=0,"—",COUNTIF({items_range_p},1)/COUNT({items_range_p}))', '0.000'),
    ('Mean κ* across items',                    f'=IFERROR(AVERAGE({items_range_t}),"—")',                       '0.000'),
    ('',                                        None, None),
    ('Items PASS (I-CVI ≥ pass threshold)',     f'=COUNTIF({items_range_u},"PASS")',         '0'),
    ('Items BORDERLINE',                         f'=COUNTIF({items_range_u},"BORDERLINE")',   '0'),
    ('Items FAIL',                               f'=COUNTIF({items_range_u},"FAIL")',         '0'),
]

row = 4
cell(sm, f'A{row}', 'HEADLINE METRICS', bold=True, size=13, color='E27431')
row += 1
for label, formula, fmt in metrics:
    if not label:
        row += 1
        continue
    cell(sm, f'A{row}', label, size=11, border=True, bg='F8F4ED')
    if formula is not None:
        c = cell(sm, f'B{row}', formula, size=11, bold=True, border=True, h='center', fmt=fmt)
    row += 1

# Highlight key cells
# S-CVI/Ave conditional
sm.conditional_formatting.add('B8', FormulaRule(formula=['AND(ISNUMBER(B8),B8>=0.9)'],  fill=green, font=Font(name=FONT, bold=True, color='008000')))
sm.conditional_formatting.add('B8', FormulaRule(formula=['AND(ISNUMBER(B8),B8>=0.8,B8<0.9)'], fill=amber, font=Font(name=FONT, bold=True, color='C77800')))
sm.conditional_formatting.add('B8', FormulaRule(formula=['AND(ISNUMBER(B8),B8<0.8)'],  fill=red,   font=Font(name=FONT, bold=True, color='CC0000')))

sm.conditional_formatting.add('B9', FormulaRule(formula=['AND(ISNUMBER(B9),B9>=0.8)'],  fill=green, font=Font(name=FONT, bold=True, color='008000')))
sm.conditional_formatting.add('B9', FormulaRule(formula=['AND(ISNUMBER(B9),B9<0.8)'],   fill=amber, font=Font(name=FONT, bold=True, color='C77800')))

sm.conditional_formatting.add('B10', FormulaRule(formula=['AND(ISNUMBER(B10),B10>=0.74)'], fill=green, font=Font(name=FONT, bold=True, color='008000')))
sm.conditional_formatting.add('B10', FormulaRule(formula=['AND(ISNUMBER(B10),B10>=0.60,B10<0.74)'], fill=amber, font=Font(name=FONT, bold=True, color='C77800')))
sm.conditional_formatting.add('B10', FormulaRule(formula=['AND(ISNUMBER(B10),B10<0.60)'], fill=red, font=Font(name=FONT, bold=True, color='CC0000')))

# ── Per-Type Breakdown ──────────────────────────────────────────────
row += 1
cell(sm, f'A{row}', 'BREAKDOWN BY CONTENT TYPE', bold=True, size=13, color='E27431')
row += 1

# Header
cell(sm, f'A{row}', 'Content Type',  bold=True, size=11, color='FFFFFF', bg='2D5A3D', h='center', border=True)
cell(sm, f'B{row}', 'Items',          bold=True, size=11, color='FFFFFF', bg='2D5A3D', h='center', border=True)
cell(sm, f'C{row}', 'S-CVI/Ave (this type)', bold=True, size=11, color='FFFFFF', bg='2D5A3D', h='center', border=True)
cell(sm, f'D{row}', 'PASS',           bold=True, size=11, color='FFFFFF', bg='2D5A3D', h='center', border=True)
row += 1

types_to_break = ['Vocabulary', 'Lesson Sentence', 'Quiz Item', 'Story Paragraph', 'Story Vocab']
for typ in types_to_break:
    cell(sm, f'A{row}', typ, size=11, border=True, bg='F8F4ED')
    cell(sm, f'B{row}', f'=COUNTIF({items_range_b},"{typ}")', size=11, h='center', border=True, fmt='0')
    cell(sm, f'C{row}',
         f'=IFERROR(AVERAGEIFS({items_range_p},{items_range_b},"{typ}"),"—")',
         size=11, h='center', border=True, fmt='0.000', bold=True)
    cell(sm, f'D{row}',
         f'=COUNTIFS({items_range_b},"{typ}",{items_range_u},"PASS")',
         size=11, h='center', border=True, fmt='0')
    row += 1

# ── Thresholds reference ─────────────────────────────────────────────
row += 2
cell(sm, f'A{row}', 'INTERPRETATION REFERENCE', bold=True, size=13, color='E27431')
row += 1
ref = [
    ('S-CVI/Ave',  '≥ 0.90 — Excellent · 0.80–0.89 — Adequate · < 0.80 — Revise'),
    ('S-CVI/UA',   '≥ 0.80 — Strict universal-agreement target met'),
    ('Mean κ*',    '> 0.74 Excellent · 0.60–0.74 Good · 0.40–0.59 Fair · < 0.40 Poor'),
    ('I-CVI item', '≥ 0.78 (N=6–10) or ≥ 0.83 (N=3–5) — PASS'),
]
for name, desc in ref:
    cell(sm, f'A{row}', name, bold=True, size=11, border=True, bg='F8F4ED')
    cell(sm, f'B{row}', desc, size=11, border=True, wrap=True)
    sm.merge_cells(f'B{row}:D{row}')
    row += 1

# ── Save ────────────────────────────────────────────────────────────
wb.save(OUT)
print(f"\nWritten: {OUT}")
print(f"Size:    {OUT.stat().st_size:,} bytes")
